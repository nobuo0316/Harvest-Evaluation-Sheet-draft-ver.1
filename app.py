import streamlit as st
import pandas as pd
from io import BytesIO
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="情報媒体 比較表", page_icon="📊", layout="wide")

# ============================================================
# 基本データ
# ============================================================
media_data = [
    {
        "情報媒体名": "日経GX",
        "サイト名": "日経GX - 脱炭素時代における変革のヒントを伝える",
        "サイトURL": "https://project.nikkeibp.co.jp/gx/",
        "金額": "法人1人 45,606円/年",
        "比較基準": "（新エネルギー新聞 21,450円/年）",
        "媒体": "メール / アプリ / サイト",
        "主な特長": "テクノロジー・市場・政策を幅広く掲載 / 記者取材あり / サイト内検索あり / ランキングあり / 独自記事あり / 記事保存可 / 複数媒体で読める",
        "懸念点": "バイオ系のタグ付けなし（ただし検索可能） / 新エネルギー新聞より高額",
        "おすすめ度": "◎",
        "向いている使い方": "幅広い業界動向の把握、保存・共有、検索ベースの情報収集",
    },
    {
        "情報媒体名": "グリーンプロダクション",
        "サイト名": "株式会社グリーンプロダクション",
        "サイトURL": "https://greenproduction.co.jp/",
        "金額": "無料",
        "比較基準": "（新エネルギー新聞 21,450円/年）",
        "媒体": "ウェブサイトのみ",
        "主な特長": "G&B（グリーン・バイオ）分野に注力 / バイオ燃料のみ見やすい / 更新頻度は約3日に1件 / 参考URL記載あり / 展示会・イベント情報あり / 海外記事多め / 無料",
        "懸念点": "今日更新分の一覧がない / 企業・団体から掲載希望可能 / サイト内検索なし",
        "おすすめ度": "◎",
        "向いている使い方": "バイオ燃料分野を絞って追う、海外トピック確認、無料の補完情報源",
    },
    {
        "情報媒体名": "環境新聞",
        "サイト名": "環境新聞オンライン",
        "サイトURL": "https://www.kankyo-news.co.jp/",
        "金額": "オンラインのみ 26,400円/年 / 紙面付き 34,980円/年",
        "比較基準": "（新エネルギー新聞 21,450円/年）",
        "媒体": "サイト / 紙",
        "主な特長": "大判で通常6～12ページ / 月4回発行（毎週水曜） / サイト内検索あり / 日付・期間指定可 / 補助金や業務委託など公共情報あり / PDFで紙面アーカイブ確認可",
        "懸念点": "新エネルギー新聞より高い / 紙面追加で年8,580円増 / 『脱炭素・エネルギー』中心で関連の薄い情報も混ざる",
        "おすすめ度": "○",
        "向いている使い方": "紙面やPDFで一覧性を重視したい場合、公共案件情報も追いたい場合",
    },
]

comparison_df = pd.DataFrame(media_data)

# ============================================================
# 図版設定
# Streamlit Cloud では repo 内に images フォルダを作って画像を置く想定
# ============================================================
IMAGE_DIR = Path("images")

figure_data = [
    {
        "図番号": "図1",
        "タイトル": "日経GXのサイトホーム画面",
        "ファイル名": "nikkei_gx_home.png",
        "媒体": "日経GX",
        "関連URL": "https://project.nikkeibp.co.jp/gx/",
        "説明": "サイトのトップ画面。主要記事や注目コンテンツが一覧化されている。",
    },
    {
        "図番号": "図2",
        "タイトル": "日経GXサイト内で『バイオ』と検索したときの様子",
        "ファイル名": "nikkei_gx_search_bio.png",
        "媒体": "日経GX",
        "関連URL": "https://project.nikkeibp.co.jp/gx/",
        "説明": "サイト内検索により、関連キーワードの記事を横断的に確認できる。",
    },
    {
        "図番号": "図3",
        "タイトル": "グリーンプロダクションのサイトホーム画面",
        "ファイル名": "green_production_home.png",
        "媒体": "グリーンプロダクション",
        "関連URL": "https://greenproduction.co.jp/",
        "説明": "G&B分野のニュースやトピックスがカテゴリ別に掲載されている。",
    },
    {
        "図番号": "図4",
        "タイトル": "グリーンプロダクションのバイオ燃料タグの記事一例",
        "ファイル名": "green_production_article.png",
        "媒体": "グリーンプロダクション",
        "関連URL": "https://greenproduction.co.jp/",
        "説明": "記事下部に参考サイトリンクがあり、詳細情報へアクセスしやすい。",
    },
    {
        "図番号": "図5",
        "タイトル": "環境新聞のサイトホーム画面",
        "ファイル名": "kankyo_news_home.png",
        "媒体": "環境新聞",
        "関連URL": "https://www.kankyo-news.co.jp/",
        "説明": "最新記事をトップで確認できる。環境・エネルギー関連の情報が幅広い。",
    },
    {
        "図番号": "図6",
        "タイトル": "環境新聞のアーカイブPDF",
        "ファイル名": "kankyo_news_pdf.png",
        "媒体": "環境新聞",
        "関連URL": "https://www.kankyo-news.co.jp/",
        "説明": "紙面アーカイブをPDF形式で確認でき、一覧性が高い。",
    },
]

figure_df = pd.DataFrame(figure_data)

# ============================================================
# 共通関数
# ============================================================
def image_path(filename: str) -> Path:
    return IMAGE_DIR / filename


def image_exists(filename: str) -> bool:
    return image_path(filename).exists()


def style_worksheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    thin = Side(style="thin", color="CCCCCC")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def set_column_widths(ws, widths: dict):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_excel_file(filtered_df: pd.DataFrame, figure_df: pd.DataFrame) -> BytesIO:
    wb = Workbook()

    # --------------------
    # シート1：比較表
    # --------------------
    ws1 = wb.active
    ws1.title = "比較表"

    export_df = filtered_df[[
        "情報媒体名", "サイト名", "サイトURL", "金額", "媒体", "おすすめ度",
        "向いている使い方", "主な特長", "懸念点"
    ]].copy()

    ws1.append(list(export_df.columns))
    for row in export_df.itertuples(index=False):
        ws1.append(list(row))

    style_worksheet(ws1)
    set_column_widths(ws1, {
        1: 18, 2: 34, 3: 38, 4: 28, 5: 18, 6: 10, 7: 28, 8: 55, 9: 45
    })
    ws1.freeze_panes = "A2"

    for row_num in range(2, ws1.max_row + 1):
        url_cell = ws1.cell(row=row_num, column=3)
        if url_cell.value:
            url_cell.hyperlink = str(url_cell.value)
            url_cell.style = "Hyperlink"

    # --------------------
    # シート2：図版一覧
    # --------------------
    ws2 = wb.create_sheet(title="図版一覧")
    headers = ["図番号", "タイトル", "媒体", "関連URL", "説明", "画像"]
    ws2.append(headers)
    style_worksheet(ws2)
    set_column_widths(ws2, {1: 10, 2: 40, 3: 20, 4: 38, 5: 35, 6: 18})
    ws2.freeze_panes = "A2"

    current_row = 2
    for item in figure_df.to_dict(orient="records"):
        ws2.cell(row=current_row, column=1, value=item["図番号"])
        ws2.cell(row=current_row, column=2, value=item["タイトル"])
        ws2.cell(row=current_row, column=3, value=item["媒体"])
        ws2.cell(row=current_row, column=4, value=item["関連URL"])
        ws2.cell(row=current_row, column=5, value=item["説明"])
        ws2.cell(row=current_row, column=4).hyperlink = item["関連URL"]
        ws2.cell(row=current_row, column=4).style = "Hyperlink"

        img_file = image_path(item["ファイル名"])
        if img_file.exists():
            img = XLImage(str(img_file))
            img.width = 320
            img.height = 180
            ws2.add_image(img, f"F{current_row}")
            ws2.row_dimensions[current_row].height = 145
        else:
            ws2.cell(row=current_row, column=6, value="画像ファイル未配置")

        current_row += 1

    # --------------------
    # シート3：結論
    # --------------------
    ws3 = wb.create_sheet(title="結論")
    ws3["A1"] = "結論"
    ws3["A1"].font = Font(bold=True, size=14)
    ws3["A3"] = "『日経GX』と『グリーンプロダクション』の併用が、バイオマス事業の情報収集に適していると考える。"
    ws3["A5"] = "理由"
    ws3["A5"].font = Font(bold=True)
    ws3["A6"] = "・日経GX：記事数が多く、検索性・保存性に優れる。"
    ws3["A7"] = "・グリーンプロダクション：無料で、海外・バイオ燃料分野の補完に向く。"
    ws3["A8"] = "・環境新聞：紙面/PDFの一覧性は高いが、費用面と情報の広さの面で優先度は一段下がる。"
    ws3["A10"] = "検索結果メモ"
    ws3["A10"].font = Font(bold=True)
    ws3["A11"] = "・『バイオマス発電』の1年分検索結果：環境新聞 19件、日経GX 31件"
    ws3["A13"] = "懸念点"
    ws3["A13"].font = Font(bold=True)
    ws3["A14"] = "・日経GXは45,606円/年で、新エネルギー新聞21,450円/年の約2.13倍。"
    ws3["A15"] = "・日経GXとグリーンプロダクションは紙面/PDFのような一覧性が弱い。"
    ws3.column_dimensions["A"].width = 95
    for r in range(1, 16):
        ws3[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================================
# 画面ヘッダー
# ============================================================
st.title("情報媒体 比較表")
st.caption("バイオマス事業の情報収集向けに、候補媒体を一目で比較できる表と参考図版をまとめています。")

col1, col2, col3 = st.columns(3)
with col1:
    st.metric("有料媒体数", "2")
with col2:
    st.metric("無料媒体数", "1")
with col3:
    st.metric("推奨構成", "日経GX + グリーンプロダクション")

st.markdown("---")

# ============================================================
# フィルター
# ============================================================
left, right = st.columns([1, 2])
with left:
    selected_media = st.multiselect(
        "表示する媒体を選択",
        options=comparison_df["情報媒体名"].tolist(),
        default=comparison_df["情報媒体名"].tolist(),
    )

with right:
    keyword = st.text_input("キーワード検索（特長・懸念点・使い方）", placeholder="例：検索、海外、無料、PDF")

filtered_df = comparison_df.copy()
filtered_df = filtered_df[filtered_df["情報媒体名"].isin(selected_media)]

if keyword:
    mask = filtered_df.apply(
        lambda row: row.astype(str).str.contains(keyword, case=False, na=False).any(), axis=1
    )
    filtered_df = filtered_df[mask]

filtered_display = filtered_df.copy()
filtered_display["サイト"] = filtered_display.apply(
    lambda row: f'<a href="{row["サイトURL"]}" target="_blank">{row["サイト名"]}</a>', axis=1
)
filtered_display = filtered_display[
    [
        "情報媒体名",
        "サイト",
        "金額",
        "媒体",
        "おすすめ度",
        "向いている使い方",
        "主な特長",
        "懸念点",
    ]
]

# ============================================================
# 比較表
# ============================================================
st.subheader("比較一覧")
st.markdown(filtered_display.to_html(escape=False, index=False), unsafe_allow_html=True)
st.info("サイト名クリックで対象サイトへ移動できます。")

# ============================================================
# 図版セクション
# ============================================================
st.subheader("参考図版")
st.caption("各媒体のホーム画面や記事例、PDFアーカイブ例を一覧で確認できます。")

figure_tabs = st.tabs([item["図番号"] for item in figure_data])
for tab, item in zip(figure_tabs, figure_data):
    with tab:
        st.markdown(f"**{item['図番号']}：{item['タイトル']}**")
        st.markdown(f"媒体：{item['媒体']}  ")
        st.markdown(f"サイト：[{item['関連URL']}]({item['関連URL']})")
        st.write(item["説明"])

        img_file = image_path(item["ファイル名"])
        if img_file.exists():
            st.image(str(img_file), use_container_width=True)
        else:
            st.warning(f"画像ファイルが見つかりません: {img_file}")
            st.code(f"{img_file}")

st.markdown("---")

# ============================================================
# 結論
# ============================================================
st.subheader("結論")
st.markdown(
    """
**『日経GX』と『グリーンプロダクション』の併用が適している** と考えられます。

**理由**
- **日経GX**：記事数が多く、サイト内検索や保存機能があり、情報の深さと使いやすさに優れる。
- **グリーンプロダクション**：無料で、バイオ燃料系や海外トピックを補完しやすい。
- **環境新聞**：紙面・PDFの一覧性は魅力だが、費用と情報の広さの面で優先度は一段下がる。

**検索結果メモ**
- 「バイオマス発電」で1年分検索した場合：
    - 環境新聞：19件
    - 日経GX：31件

**主な懸念点**
- 日経GXは **45,606円/年** と、新エネルギー新聞 **21,450円/年** の約2.13倍。
- 日経GX・グリーンプロダクションともに、紙面やPDFのような一覧性は弱く、記事ごとの確認が必要。
"""
)

# ============================================================
# Excel出力
# ============================================================
st.subheader("Excel出力")
st.write("比較表・図版一覧・結論をまとめたExcelをダウンロードできます。")

excel_file = build_excel_file(filtered_df, figure_df)
st.download_button(
    label="Excelをダウンロード",
    data=excel_file,
    file_name="media_comparison_with_figures.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

# ============================================================
# 元データ確認
# ============================================================
with st.expander("元データを確認"):
    st.dataframe(comparison_df, use_container_width=True, hide_index=True)
    st.dataframe(figure_df, use_container_width=True, hide_index=True)

# ============================================================
# 画像配置ガイド
# ============================================================
with st.expander("画像ファイルの置き方"):
    st.markdown(
        """
リポジトリ内に **images** フォルダを作り、下記ファイル名で保存してください。

- `images/nikkei_gx_home.png`
- `images/nikkei_gx_search_bio.png`
- `images/green_production_home.png`
- `images/green_production_article.png`
- `images/kankyo_news_home.png`
- `images/kankyo_news_pdf.png`

そのまま配置すれば、Streamlit画面表示とExcel出力の両方で使えます。
"""
    )

st.markdown("---")
st.caption("必要であれば次に、①稟議用の落ち着いた配色版 ②画像を縦並びレポート風にした版 ③CSV管理対応版 に整えられます。")
